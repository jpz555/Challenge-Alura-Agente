"""
=========================================================
Document Processor - Parser
=========================================================
Responsabilidad:
    Leer documentos y devolver su contenido.

Versión:
    1.0.0
=========================================================
"""

from pathlib import Path
from typing import Dict, List

from document_processor.config import SUPPORTED_FORMATS, ENCODING
from openpyxl import load_workbook


class DocumentParser:
    """
    Parser principal del Document Processor.

    En esta primera versión únicamente soporta Markdown.
    Posteriormente se agregarán PDF, DOCX, PPTX, XLSX, etc.
    """

    def __init__(self):
        pass

    # =====================================================
    # Lectura de un documento
    # =====================================================

    def read(self, file_path: Path) -> Dict:
        """
        Lee un documento y devuelve un diccionario.

        Parameters
        ----------
        file_path : Path

        Returns
        -------
        dict
        """

        extension = file_path.suffix.lower()

        if extension not in SUPPORTED_FORMATS:
            raise ValueError(
                f"Formato no soportado: {extension}"
            )

        if extension == ".md":
            return self._read_markdown(file_path)
        
        if extension == ".xlsx":
            return self._read_excel(file_path)

        raise NotImplementedError(
            f"Parser para {extension} aún no implementado."
        )

    # =====================================================
    # Markdown
    # =====================================================
    def _read_markdown(self, file_path: Path) -> Dict:

        with open(file_path,"r",encoding=ENCODING) as f:

            text = f.read()

        return {
            "file_name": file_path.name,
            "file_path": str(file_path),
            "category": file_path.parent.name,
            "extension": ".md",
            "content": text
        }
        
    # Excel
    def _read_excel(self,file_path: Path,) -> Dict:
        """
        Lee un archivo Excel y convierte cada hoja en texto
        estructurado para el proceso de chunking.
        """
        workbook = load_workbook(filename=file_path,data_only=True)

        sections = []
        
        for sheet in workbook.worksheets:
            
            sections.append(f"# Hoja: {sheet.title}")
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = ["" if h is None else str(h)for h in rows[0]]

            for row in rows[1:]:
                record = []
                for header, value in zip(headers, row):
                    value = "" if value is None else str(value)

                    record.append(f"{header}: {value}")

                sections.append("\n".join(record))
                sections.append("")

        text = "\n".join(sections)

        return {
            "file_name": file_path.name,
            "file_path": str(file_path),
            "category": file_path.parent.name,
            "extension": ".xlsx",
            "content": text
        }

    # =====================================================
    # Lectura masiva
    # =====================================================
    def read_directory(self, directory: Path) -> List[Dict]:
        """
    Lee recursivamente todos los documentos soportados
    dentro de una carpeta y sus subcarpetas.
    """

        documents = []

        # Buscar de forma recursiva
        for file in sorted(directory.rglob("*")):

            if (
                file.is_file()
                and file.suffix.lower() in SUPPORTED_FORMATS
            ):

                documents.append(
                    self.read(file)
                )

        return documents